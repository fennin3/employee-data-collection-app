from os import error
from django.http import request
from django.shortcuts import redirect, render
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from .forms import EmployeeCreation
from django.contrib import messages
from .models import Employee, Supervisor, Log
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


@login_required
def home(request):
    employees = reversed(Employee.objects.all())
    context = {"employees": employees}
    return render(request, "base/index.html", context)


@login_required
def log(request):
    logs = reversed(Log.objects.all())
    context = {"logs": logs}
    return render(request, "base/logs.html", context)


def user_login(request):
    if request.method == "POST":
        username = request.POST.get("username", "")
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user=user)
            return redirect("home")
        else:
            return render(request, "base/login.html")
    return render(request, "base/login.html")

@login_required
def employee_creation(request):
    if request.method == "POST":
        form = EmployeeCreation(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee added successfully")
            return redirect("home")
        else:
            messages.error(request, "Please check input")
            form = EmployeeCreation()
            return render(request, "base/create_employee.html", {"form": form})
    else:
        form = EmployeeCreation()
        return render(request, "base/create_employee.html", {"form": form})

@login_required
def upload_data(request):
    if request.method == "POST":
        try:
            file_ = request.FILES["file"]
        except Exception as e:
            file_ = ""
        filename = str(file_)
        ext = filename.split(".")[-1]
        sup = Supervisor.objects.all()
        sub2 = []
        print(ext)
        if ext == "xlsx":
            df1=pd.read_excel(file_,engine='openpyxl',)
            data = df1.drop_duplicates(subset=['first_name','last_name', 'Age','Salary'])
            print(data)
            wb = openpyxl.Workbook()
            ws = wb.active
            for r in dataframe_to_rows(data, index=False, header=False):
                ws.append(r)

            print(ws.max_column)
            # wb = openpyxl.load_workbook(data)
            # wb = wb.active
            # print(wb)

            try:
                for i in range(1, ws.max_row + 1):
                    for j in sup:
                        if j.firstname in str(ws.cell(row=i, column=9).value):
                            sub2.append(j)

                    print(f"List is ===> {sub2}")

                    employee = Employee.objects.create(
                        first_name=ws.cell(row=i, column=1).value,
                        last_name=ws.cell(row=i, column=2).value,
                        age=ws.cell(row=i, column=3).value,
                        date_of_birth=ws.cell(row=i, column=4).value,
                        date_of_employment=ws.cell(row=i, column=5).value,
                        position=ws.cell(row=i, column=6).value,
                        department=ws.cell(row=i, column=7).value,
                        salary=ws.cell(row=i, column=8).value,
                    )
                    employee.save()

                    for i in sub2:
                        print(i)
                        employee.supervisors.add(i)
                    employee.save()
                    sub2 = []
                log = Log.objects.create(
                    number_of_employee_data=ws.max_row - 1,
                    status="Successful",
                    error="No errors",
                )
                log.save()
            except Exception as e:
                log = Log.objects.create(
                    number_of_employee_data=ws.max_row,
                    status="Unsuccessful",
                    error=e,
                )
                log.save()

            return redirect("home")
        elif ext == "":
            messages.error(request, "Please choose a file")
        else:
            messages.error(
                request, "Please upload an Excel file with '.xlsx' extension"
            )
            return redirect("upload_data")
    return render(request, "base/upload.html")
