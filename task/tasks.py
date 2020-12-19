from __future__ import absolute_import, unicode_literals
from django.shortcuts import redirect, render
from openpyxl.utils.dataframe import dataframe_to_rows
from django.contrib import messages

from django.contrib.auth.decorators import login_required
from base.models import Log, Supervisor, Employee
from celery import shared_task
import pandas as pd
import openpyxl

message_ = ""


@shared_task
def upload_task(file_):
    sup = Supervisor.objects.all()
    sub2 = []
    df1 = pd.read_excel(
        file_,
        engine="openpyxl",
    )
    data = df1.drop_duplicates(subset=["first_name", "last_name", "Age", "Salary"])
    print(data)
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(data, index=False, header=False):
        ws.append(r)

    print(ws.max_column)
    # wb = openpyxl.load_workbook(data)
    # wb = wb.active
    # print(wb)

    try:
        for i in range(1, ws.max_row + 1):
            for j in sup:
                if j.firstname in str(ws.cell(row=i, column=9).value):
                    sub2.append(j)

            print(f"List is ===> {sub2}")

            employee = Employee.objects.create(
                first_name=ws.cell(row=i, column=1).value,
                last_name=ws.cell(row=i, column=2).value,
                age=ws.cell(row=i, column=3).value,
                date_of_birth=ws.cell(row=i, column=4).value,
                date_of_employment=ws.cell(row=i, column=5).value,
                position=ws.cell(row=i, column=6).value,
                department=ws.cell(row=i, column=7).value,
                salary=ws.cell(row=i, column=8).value,
            )
            employee.save()

            for i in sub2:
                print(i)
                employee.supervisors.add(i)
            employee.save()
            sub2 = []
        log = Log.objects.create(
            number_of_employee_data=ws.max_row,
            status="Successful",
            error="No errors",
        )
        log.save()

    except Exception as e:
        log = Log.objects.create(
            number_of_employee_data=ws.max_row,
            status="Unsuccessful",
            error=e,
        )
        log.save()
